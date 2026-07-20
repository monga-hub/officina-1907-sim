#!/usr/bin/env python3
"""Estrae i dati di Officina 1907 da 'mazzi - ordinato.xlsx' e genera src/game/data.js"""
import openpyxl, json, re, sys

XLSX = '/Users/daniele/Desktop/Giochi da tavolo/OFFICINA 1907/plance/mazzi - ordinato.xlsx'
OUT = '/Users/daniele/CoudeCode/officina-1907-sim/src/game/data.js'

wb = openpyxl.load_workbook(XLSX, data_only=True)

NATIONS = ['Italiani', 'Francesi', 'Polacchi', 'Spagnoli', 'Tedeschi']
SECTORS = {'Chimica', 'Metallurgica', 'Tessile'}

def norm(s):
    return re.sub(r'\s+', ' ', str(s or '')).strip()

def parse_effect(text, icona, sector):
    t = norm(text).lower()
    if 'per 3 monete' in t or 'per 3 monete' in t.replace('risorsaper', 'risorsa per'):
        return {'type': 'swap_res_3m'}
    if t.startswith('scambia una risorsa'):
        return {'type': 'swap_res_any'}
    if t.startswith('scambia 2 monete') or t.startswith('scambia due monete'):
        return {'type': 'buy_res_2m'}
    if 'moneta per ogni nazione diversa' in t:
        return {'type': 'coin_per_nation', 'amount': 1}
    if t == 'prendi questa risorsa':
        return {'type': 'take_res'}
    if t == 'prendi tre monete':
        return {'type': 'take_coins', 'amount': 3}
    if 'monete per ogni icona' in t:
        ic = norm(icona)
        m = re.match(r'Icona settore (\w+)', ic)
        if m:
            return {'type': 'coins_per_icon', 'amount': 2, 'icon': {'kind': 'sector', 'value': m.group(1)}}
        m = re.match(r'Icona nazionalit\S+ (\w+)', ic)
        if m:
            return {'type': 'coins_per_icon', 'amount': 2, 'icon': {'kind': 'nation', 'value': m.group(1)}}
        raise ValueError(f'icona non riconosciuta: {ic!r} per effetto {text!r}')
    if 'risorsa per livello di tensione' in t:
        return {'type': 'res_per_tension'}
    raise ValueError(f'effetto non riconosciuto: {text!r}')

workers = []
wid = 0
for nation in NATIONS:
    ws = wb[nation]
    rows = list(ws.iter_rows(values_only=True))
    header = rows[0]
    for row in rows[1:]:
        if not any(c is not None for c in row):
            continue
        copie, naz, settore, v, effetto, icona = row[0], norm(row[1]), norm(row[2]), row[3], row[4], row[5]
        assert naz == nation, f'{naz} != {nation}'
        assert settore in SECTORS, settore
        eff = parse_effect(effetto, icona, settore)
        for _ in range(int(copie)):
            wid += 1
            workers.append({
                'id': f'w{wid}', 'nation': nation, 'sector': settore, 'v': int(v),
                'effect': eff, 'effectText': norm(effetto),
            })

assert len(workers) == 75, len(workers)

# Welfare/Macchinari: nomi dal regolamento giocatore, mappati per (V, S1, S2)
WELFARE_NAMES = {
    (5, 'Tessile', 'Metallurgica'): 'Mensa aziendale',
    (5, 'Tessile', 'Chimica'): 'Spaccio aziendale',
    (5, 'Chimica', 'Tessile'): 'Asilo operaio',
    (5, 'Chimica', 'Metallurgica'): 'Casa di ringhiera',
    (5, 'Metallurgica', 'Chimica'): 'Trasporto operaio',
    (5, 'Metallurgica', 'Tessile'): 'Infermeria di fabbrica',
    (7, 'Tessile', 'Metallurgica'): 'Colonia estiva',
    (7, 'Tessile', 'Chimica'): 'Biblioteca popolare',
    (7, 'Chimica', 'Metallurgica'): 'Camera del lavoro',
    (7, 'Chimica', 'Tessile'): 'Squadra di fabbrica',
    (7, 'Metallurgica', 'Tessile'): 'Orti operai',
    (7, 'Metallurgica', 'Chimica'): 'Banda musicale',
}
welfare = []
ws = wb['Welfare e Macchinari']
i = 0
for row in list(ws.iter_rows(values_only=True))[1:]:
    if not any(c is not None for c in row):
        continue
    _, s1, s2, v, t1, t2, eff, nota = [norm(c) if isinstance(c, str) else c for c in row[:8]]
    v, t1, t2 = int(v), int(t1), int(t2)
    i += 1
    welfare.append({
        'id': f'wf{i}', 'name': WELFARE_NAMES[(v, s1, s2)],
        'v': v, 's1': s1, 's2': s2, 't1': t1, 't2': t2,
        # Macchinario (Sotto): produzione a inizio turno.
        # V=5: 1 risorsa del settore 1°, max 3 usi. V=7: 1 risorsa per ciascun settore, max 2 usi.
        'usesMax': 3 if v == 5 else 2,
        'perUse': [s1] if v == 5 else [s1, s2],
    })
assert len(welfare) == 12

def parse_reslist(s):
    parts = [norm(p) for p in str(s).split(',')]
    for p in parts:
        assert p in SECTORS, p
    return parts

contracts = {}
for size, sheet, pv in [('small', 'Commesse piccole', [5, 3]), ('medium', 'Commesse medie', [9, 7]), ('large', 'Commesse grandi', [13, 10])]:
    cards = []
    ws = wb[sheet]
    for j, row in enumerate(list(ws.iter_rows(values_only=True))[1:]):
        if not any(c is not None for c in row):
            continue
        cards.append({
            'id': f'{size}{j+1}', 'size': size, 'pv': pv,
            'reqs': [parse_reslist(row[1]), parse_reslist(row[2])],
        })
    assert len(cards) == 6, (size, len(cards))
    contracts[size] = cards

# Obiettivi (Piano Industriale)
def parse_obj(text):
    t = norm(text)
    m = re.match(r'Tracciati (\w+) e (\w+) sviluppati fino alla milestone', t)
    if m:
        return {'type': 'milestones', 'sectors': [m.group(1), m.group(2)]}
    m = re.match(r'(\d) lavoratori (\w+) installati in fabbrica', t)
    if m and m.group(2) in NATIONS:
        return {'type': 'workers_nation', 'n': int(m.group(1)), 'nation': m.group(2)}
    m = re.match(r'(\d) lavoratori della stessa nazionalit', t)
    if m:
        return {'type': 'same_nation', 'n': int(m.group(1))}
    m = re.match(r'(\d) lavoratori di nazionalit\S+ diverse', t)
    if m:
        return {'type': 'distinct_nations', 'n': int(m.group(1))}
    if 'Tracciati Tensione a 0 contemporaneamente' in t:
        return {'type': 'all_tension_zero'}
    m = re.search(r'guadagni almeno (\d+) marchi', t)
    if m:
        return {'type': 'activation_coins', 'n': int(m.group(1))}
    m = re.match(r'Almeno (\d) cart[ae] Sotto installat[ae] in ciascuno', t)
    if m:
        return {'type': 'sotto_each', 'n': int(m.group(1))}
    m = re.match(r'Almeno (\d) carte Sopra installate in ciascuno', t)
    if m:
        return {'type': 'sopra_each', 'n': int(m.group(1))}
    if 'fine partita nessuna carta bloccata' in t:
        return {'type': 'no_blocked_end'}
    m = re.match(r'(\d) carte Welfare \(posizione Sopra\)', t)
    if m:
        return {'type': 'direzione', 'side': 'sopra', 'n': int(m.group(1))}
    m = re.match(r'(\d) carte Macchinario \(posizione Sotto\)', t)
    if m:
        return {'type': 'direzione', 'side': 'sotto', 'n': int(m.group(1))}
    m = re.match(r'(\d) carte \(Welfare o Macchinario', t)
    if m:
        return {'type': 'direzione', 'side': 'any', 'n': int(m.group(1))}
    if re.match(r'In un reparto a scelta: almeno 3 carte Sopra e 2 carte Sotto', t):
        return {'type': 'full_dept', 'sopra': 3, 'sotto': 2}
    raise ValueError(f'obiettivo non riconosciuto: {t!r}')

objectives = []
ws = wb['Obiettivi Aziendali']
for row in list(ws.iter_rows(values_only=True))[1:]:
    if not any(c is not None for c in row):
        continue
    tid = norm(row[0]).split(' ')[1]
    objs = []
    for text, pv in [(row[1], row[2]), (row[3], row[4]), (row[5], row[6])]:
        objs.append({'text': norm(text), 'pv': int(pv), 'cond': parse_obj(text)})
    objectives.append({'id': f't{tid}', 'name': f'Tessera {tid}', 'objectives': objs})
assert len(objectives) == 32, len(objectives)

# ---- Dati statici (regolamento giocatore + plance 26 + design doc) ----
static_js = """
// Settori e risorse
export const SECTORS = ['Tessile', 'Metallurgica', 'Chimica'];
export const RESOURCE_OF = { Tessile: 'Tessuti', Metallurgica: 'Acciaio', Chimica: 'Coloranti' };
export const SECTOR_COLORS = { Tessile: '#b0413e', Metallurgica: '#3e6f8f', Chimica: '#b09c3e' };
export const NATIONS = ['Italiani', 'Francesi', 'Polacchi', 'Spagnoli', 'Tedeschi'];
export const NATION_FLAGS = { Italiani: '🇮🇹', Francesi: '🇫🇷', Polacchi: '🇵🇱', Spagnoli: '🇪🇸', Tedeschi: '🇩🇪' };

// Plancia centrale: 5 nodi perimetrali + Borsa. Ogni nodo perimetrale ha 2 banchi adiacenti.
// ASSUNZIONE: abbinamento banchi-nodi non specificato nei file — layout fisso a pentagono.
export const NODES = ['Tessile', 'Metallurgica', 'Chimica', 'Servizi', 'Sindacato', 'Borsa'];
export const NODE_BANKS = {
  Tessile: ['Francesi', 'Italiani'],
  Metallurgica: ['Italiani', 'Polacchi'],
  Chimica: ['Polacchi', 'Spagnoli'],
  Servizi: ['Spagnoli', 'Tedeschi'],
  Sindacato: ['Tedeschi', 'Francesi'],
};

// Le 6 plance Fabbrica da "Plance Fabbrica.xlsx" (terziario sx 3 slot Sopra / secondario 4 / primario 4; Sotto max 2 ovunque)
export const BOARDS = [
  { id: 'p1', name: 'Plancia 1', terziario: 'Tessile', secondario: 'Chimica', primario: 'Metallurgica' },
  { id: 'p2', name: 'Plancia 2', terziario: 'Tessile', secondario: 'Metallurgica', primario: 'Chimica' },
  { id: 'p3', name: 'Plancia 3', terziario: 'Metallurgica', secondario: 'Tessile', primario: 'Chimica' },
  { id: 'p4', name: 'Plancia 4', terziario: 'Chimica', secondario: 'Tessile', primario: 'Metallurgica' },
  { id: 'p5', name: 'Plancia 5', terziario: 'Chimica', secondario: 'Metallurgica', primario: 'Tessile' },
  { id: 'p6', name: 'Plancia 6', terziario: 'Metallurgica', secondario: 'Chimica', primario: 'Tessile' },
];
// Cap fisso per OGNI reparto: 3 Sopra + 2 Sotto, carte bloccate da Sciopero incluse (conferma autore 07/07/2026)
export const ROLE_SLOTS_SOPRA = { terziario: 3, secondario: 3, primario: 3 };

// Tracciato Produzione da "Plance Fabbrica.xlsx": griglia 4x4 (righe A alto → D basso),
// serpentina da D1: D1→D4, C4→C1, B1→B4, A4→A1 = posizioni 1..16 (confermato dall'autore).
// Semantica (confermata dall'autore): TUTTE le caselle raggiunte riproducono a ogni attivazione
// (marchi fissi, risorse, "1 marco per carta [Settore]"); le stesse caselle si incassano anche
// una tantum all'attraversamento (regolamento: "bonus di casella attraversata").
// Caselle PV = soglie di fine partita (cumulative). Milestone = solo per Obiettivi.
// cell: null | {coins:n} | {res:1} | {coinsPerIcon:1} | {pv:n} | {milestone:true}
// grid: [riga A|B|C|D][colonna 1..4] per il rendering fedele alla plancia fisica.
// Variante del 07/07/2026 (richiesta autore): «1 marco per carta [Settore]» sostituito
// ovunque da «1 marco»; casella C4 = 1 risorsa del reparto invece di 1 marco.
const T = null;
export const TRACKS = {
  terziario: [
    null,                                            // pos 0 inutilizzata (si parte da D1=1)
    { coins: 1 }, T, T, T,                           // D1..D4
    { res: 1 }, T, { pv: 2 }, T,                     // C4..C1
    { coins: 1 }, T, { res: 1 }, { milestone: true },// B1..B4
    T, T, { pv: 3 }, { coins: 1 },                   // A4..A1
  ],
  secondario: [
    null,
    { coins: 1 }, T, T, T,                           // D1..D4
    { res: 1 }, T, { pv: 2 }, T,                     // C4..C1
    { res: 1 }, T, { coins: 1 }, T,                  // B1..B4
    { milestone: true }, { pv: 3 }, T, { coins: 2 }, // A4..A1
  ],
};
TRACKS.primario = TRACKS.secondario;
export const TRACK_MAX = 16;
export const MILESTONE_POS = { terziario: 12, secondario: 13, primario: 13 };
// mappa posizione → [riga, colonna] della griglia 4x4 (riga 0=A ... 3=D) per la UI
export function trackGridPos(pos) {
  if (pos <= 4) return [3, pos - 1];        // D1..D4
  if (pos <= 8) return [2, 8 - pos];        // C4..C1
  if (pos <= 12) return [1, pos - 9];       // B1..B4
  return [0, 16 - pos];                     // A4..A1
}

// Setup (confermato dall'autore): segnalini Produzione tutti su D1 (pos 1);
// Tensione: terziario 0, secondario 1, primario 1. Nessuna risorsa iniziale.
export const SETUP = { primario: { prod: 1, tension: 1 }, secondario: { prod: 1, tension: 1 }, terziario: { prod: 1, tension: 0 } };

// Marchi iniziali per posizione di turno (1°/2°: 10, 3°/4°: 11)
export const STARTING_COINS = [10, 10, 11, 11];

export const TENSION_LIMIT = 3;
export const CLOCK_THRESHOLD = { 2: 8, 3: 12, 4: 16 };
export const CLOCK_REFRESH = [3, 6, 9, 12, 15];
export const MOVE_COSTS = [0, 1]; // slot 1 gratis, slot 2 = 1 marco
export const MAX_CONTRACTS_PER_VISIT = 2;
export const DIREZIONE_MAX = { sopra: 2, sotto: 2 };
export const UNBLOCK_COST = 3;
"""

with open(OUT, 'w') as f:
    f.write('// GENERATO da extract_data.py — dati da "mazzi - ordinato.xlsx" + regolamento giocatore\n')
    f.write(static_js)
    f.write('\nexport const WORKERS = ' + json.dumps(workers, ensure_ascii=False, indent=1) + ';\n')
    f.write('\nexport const WELFARE = ' + json.dumps(welfare, ensure_ascii=False, indent=1) + ';\n')
    f.write('\nexport const CONTRACTS = ' + json.dumps(contracts, ensure_ascii=False, indent=1) + ';\n')
    f.write('\nexport const OBJECTIVE_TILES = ' + json.dumps(objectives, ensure_ascii=False, indent=1) + ';\n')

# report riassuntivo
from collections import Counter
cs = Counter((w['sector']) for w in workers)
cv = Counter((w['v']) for w in workers)
ce = Counter(w['effect']['type'] for w in workers)
print('75 operai OK. Settori:', dict(cs), 'Valori:', dict(cv))
print('Effetti:', dict(ce))
print('Welfare:', len(welfare), 'Commesse:', {k: len(v) for k, v in contracts.items()}, 'Tessere:', len(objectives))
